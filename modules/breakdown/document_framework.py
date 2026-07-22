"""Reusable definitions and editorial pages for CinePlan production documents."""

from dataclasses import dataclass
from collections.abc import Mapping
import re

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from modules.breakdown.document_layout import THEME, editorial_pdf


@dataclass(frozen=True)
class SceneColorIdentity:
    label: str
    hex: str
    assigned: bool


_STRIPBOARD_COLORS = {
    "blanco": "#ffffff",
    "amarillo": "#fff3a3",
    "azul": "#b7dcff",
    "verde": "#c8e6c9",
    "rosa": "#f8bbd0",
    "morado": "#d1c4e9",
}
NEUTRAL_SCENE_COLOR = "#d1d5db"


def scene_color_identity(scene):
    """Resolve the existing Stripboard Color once for every document consumer."""
    source = scene if isinstance(scene, Mapping) else {}
    raw = source.get("Color stripboard", source.get("color_stripboard", ""))
    label = str(raw or "").strip()
    if re.fullmatch(r"#[0-9a-fA-F]{6}", label):
        return SceneColorIdentity(label, label.lower(), True)
    resolved = _STRIPBOARD_COLORS.get(label.casefold())
    if resolved:
        return SceneColorIdentity(label, resolved, True)
    return SceneColorIdentity("Neutral", NEUTRAL_SCENE_COLOR, False)


@dataclass(frozen=True)
class DocumentDefinition:
    label: str
    filename_title: str
    preview_generator: str
    pdf_generator: str
    excel_generator: str
    fields: tuple[str, ...] = ()
    sort_options: tuple[str, ...] = ()
    group_options: tuple[str, ...] = ()
    supports_scope: bool = False
    supports_front_matter: bool = False
    header: str = "cineplan_header"
    footer: str = "cineplan_footer"
    editorial_template: str = "cineplan_editorial"
    cover: str = "cineplan_cover"
    summary: str = "cineplan_summary"
    catalog_cover: str = "cineplan_catalog_cover"
    scene_separator: str = "cineplan_scene_separator"
    table_style: str = "cineplan_table"
    subtitle: str = "Documento de producción"
    description: str = "Documento oficial de producción CinePlan."
    metadata: tuple[str, ...] = ("Proyecto", "Versión", "Fecha", "Hora")
    color: str = "#f8b400"
    scene_color_resolver: object = scene_color_identity

    def scene_color(self, scene):
        return self.scene_color_resolver(scene)


_COMPACT_FIELDS = {"id", "cantidad", "estado", "escena", "escenas", "prioridad"}
_LONG_FIELDS = {
    "instrucción", "instruccion", "observaciones", "observación general",
    "notas", "continuidad / riesgo", "permisos", "riesgos",
}


def column_weights(columns, rows=None):
    """Balance visible columns using field semantics and representative content."""
    rows = list(rows or ())
    weights = []
    for column in columns:
        normalized = str(column).strip().casefold()
        if normalized in _COMPACT_FIELDS:
            semantic = 0.65
        elif normalized in _LONG_FIELDS:
            semantic = 2.25
        else:
            semantic = 1.15
        lengths = []
        for row in rows[:40]:
            value = row.get(column, "") if hasattr(row, "get") else ""
            lengths.append(len(str(value or "")))
        content = min(1.35, (sum(lengths) / len(lengths)) / 32) if lengths else 0
        # Numeric/identifier columns stay compact even when a stray value is long.
        weights.append(semantic if normalized in _COMPACT_FIELDS else semantic + content)
    return tuple(weights)


def pdf_column_widths(columns, usable_width, rows=None):
    weights = column_weights(columns, rows)
    total = sum(weights) or 1
    return [usable_width * weight / total for weight in weights]


def format_excel_worksheet(worksheet):
    """Apply readable widths and wrapped text without changing cell values."""
    headers = [cell.value or "" for cell in worksheet[1]] if worksheet.max_row else []
    weights = column_weights(headers)
    for index, (header, weight) in enumerate(zip(headers, weights), 1):
        values = [str(worksheet.cell(row, index).value or "") for row in range(1, worksheet.max_row + 1)]
        longest = max((len(line) for value in values for line in value.splitlines()), default=len(str(header)))
        width = min(60, max(9, longest + 2))
        if str(header).strip().casefold() in _LONG_FIELDS:
            width = min(60, max(28, width))
        elif weight < 1:
            width = min(14, width)
        worksheet.column_dimensions[get_column_letter(index)].width = width
        for row in range(1, worksheet.max_row + 1):
            worksheet.cell(row, index).alignment = Alignment(vertical="top", wrap_text=True)


DOCUMENT_DEFINITIONS = (
    DocumentDefinition("Hoja de Breakdown", "Breakdown", "breakdown_preview", "breakdown_pdf",
                       "breakdown_excel", supports_scope=True, supports_front_matter=True,
                       subtitle="Desglose por escena", description="Documento maestro de necesidades de producción.", color="#f8b400"),
    DocumentDefinition("Documento de Cast", "Reparto", "catalog_preview", "catalog_pdf", "catalog_excel",
                       ("ID", "Personaje", "Actor / Actriz", "Contacto", "Escenas", "Apariciones", "Notas"),
                       ("ID", "Personaje", "Apariciones"),
                       subtitle="Catálogo general", description="Personajes, talento y apariciones del proyecto.", color="#a855f7"),
    DocumentDefinition("Documento de Props y Utilería", "Utilería", "catalog_preview", "catalog_pdf", "catalog_excel",
                       ("Categoría", "Cantidad", "Responsable", "Personaje / Área", "Escenas", "Continuidad / Riesgo"),
                       group_options=("Categoría", "Responsable", "Escenas"),
                       subtitle="Catálogo general", description="Utilería consolidada por categoría y escena.", color="#f97316"),
    DocumentDefinition("Documento de Vestuario y Maquillaje", "Vestuario", "catalog_preview", "catalog_pdf", "catalog_excel",
                       ("Personaje", "Look", "Vestuario", "Maquillaje", "Estado", "Escenas"),
                       subtitle="Catálogo general", description="Looks, vestuario, maquillaje y continuidad.", color="#ec4899"),
    DocumentDefinition("Documento de VFX / Efectos Prácticos / Sonido", "VFX, SFX y Sonido",
                       "catalog_preview", "catalog_pdf", "catalog_excel",
                       ("Tipo", "Responsable", "Prioridad", "Seguridad", "Escenas"),
                       subtitle="Catálogo general", description="Requerimientos visuales, prácticos y sonoros.", color="#3b82f6"),
    DocumentDefinition("Documento de Extras, Vehículos y Animales", "Extras y Vehículos",
                       "catalog_preview", "catalog_pdf", "catalog_excel",
                       ("Tipo", "Responsable", "Cantidad", "Conductor", "Vehículo", "Escenas"),
                       subtitle="Catálogo general", description="Extras, vehículos, conductores y animales.", color="#10b981"),
    DocumentDefinition("Documento de Producción", "Notas de Producción", "catalog_preview", "catalog_pdf", "catalog_excel",
                       ("Departamento", "Responsable", "Instrucción", "Observaciones", "Permisos", "Riesgos"),
                       subtitle="Catálogo general", description="Instrucciones y observaciones por departamento.", color="#ef4444"),
)

DOCUMENTS_BY_LABEL = {definition.label: definition for definition in DOCUMENT_DEFINITIONS}


def _safe_project_name(value):
    cleaned = re.sub(r'[<>:"/\\|?*]+', " ", str(value or "Proyecto CinePlan"))
    return " ".join(cleaned.split()) or "Proyecto CinePlan"


def export_filename(project_name, definition, scope, scene_numbers, extension):
    project = _safe_project_name(project_name)
    suffix = definition.filename_title
    if definition.supports_scope:
        numbers = [str(number).strip() for number in scene_numbers if str(number).strip()]
        if scope == "Escena actual" and numbers:
            suffix = f"Breakdown Escena {numbers[0]}"
        elif scope == "Escenas seleccionadas":
            numeric = []
            try:
                numeric = [int(number) for number in numbers]
            except ValueError:
                pass
            consecutive = bool(numeric) and numeric == list(range(numeric[0], numeric[-1] + 1))
            joined = f"{numbers[0]}-{numbers[-1]}" if consecutive else "_".join(numbers)
            suffix = f"Breakdown Escenas {joined}"
        else:
            suffix = "Breakdown Completo"
    return f"{project} - {suffix}.{extension}"
