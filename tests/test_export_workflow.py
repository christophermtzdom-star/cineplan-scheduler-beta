import sys
import unittest
from io import BytesIO
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from streamlit.testing.v1 import AppTest

from modules.breakdown.export_page import (
    _MODE_OPTIONS,
    _export_bytes,
    _export_dataframe,
    _row_belongs_to_scenes,
    _scene_separator_pdf,
)
from modules.breakdown.document_framework import (
    DOCUMENT_DEFINITIONS,
    DOCUMENTS_BY_LABEL,
    column_weights,
    export_filename,
    pdf_column_widths,
    scene_color_identity,
)
from modules.breakdown.document_layout import (
    GRID,
    THEME,
    dataframe_pdf,
    editorial_pdf,
    html_footer,
    html_header,
    html_scene_accent,
    pdf_scene_accent_commands,
)


class ExportWorkflowTests(unittest.TestCase):
    def setUp(self):
        self.main = sys.modules["__main__"]
        self.previous = {
            name: getattr(self.main, name, None)
            for name in ("dataframe_to_pdf", "dataframe_to_excel", "generate_breakdown_pdf")
        }
        def pdf_buffer(*args, **kwargs):
            output = BytesIO()
            writer = PdfWriter()
            writer.add_blank_page(width=792, height=612)
            writer.write(output)
            output.seek(0)
            return output

        self.main.dataframe_to_pdf = pdf_buffer
        self.main.dataframe_to_excel = lambda dataframe: BytesIO(b"EXCEL")
        self.main.generate_breakdown_pdf = pdf_buffer

    def tearDown(self):
        for name, value in self.previous.items():
            if value is None:
                delattr(self.main, name)
            else:
                setattr(self.main, name, value)

    def test_all_scopes_and_formats_use_existing_generators(self):
        scenes = [
            {"Escena": "1", "Encabezado de escena": "INT. CASA"},
            {"Escena": "2", "Encabezado de escena": "EXT. CALLE"},
        ]
        scopes = ([scenes[0]], scenes[:1], scenes)
        for selected_scenes in scopes:
            self.assertTrue(_export_bytes("Hoja de Breakdown", selected_scenes, "PDF"))
            excel = _export_bytes("Hoja de Breakdown", selected_scenes, "Excel")
            self.assertTrue(excel)

    def test_full_project_repeats_the_preview_source_for_every_scene(self):
        scenes = [
            {"Escena": "1", "Encabezado de escena": "INT. CASA"},
            {"Escena": "2", "Encabezado de escena": "EXT. CALLE"},
            {"Escena": "3", "Encabezado de escena": "INT. FORO"},
        ]
        pdf = _export_bytes("Hoja de Breakdown", scenes, "PDF")
        self.assertEqual(len(PdfReader(BytesIO(pdf)).pages), 3)
        excel = _export_bytes("Hoja de Breakdown", scenes, "Excel")
        workbook = load_workbook(BytesIO(excel), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Escena 1", "Escena 2", "Escena 3"])

    def test_scene_filter_does_not_depend_on_preview_state(self):
        self.assertTrue(_row_belongs_to_scenes({"Escenas": "1, 3"}, {"3"}))
        self.assertFalse(_row_belongs_to_scenes({"Escenas": "1, 3"}, {"2"}))
        self.assertEqual(_MODE_OPTIONS, ["Escena actual", "Escenas seleccionadas", "Proyecto completo"])

    def test_breakdown_export_dataframe_contains_only_requested_scenes(self):
        scenes = [{"Escena": "7", "Encabezado de escena": "INT. FORO"}]
        dataframe = _export_dataframe("Hoja de Breakdown", scenes)
        self.assertIsInstance(dataframe, pd.DataFrame)
        self.assertEqual(set(dataframe["Escena"]), {"7"})

    def test_project_pdf_includes_optional_front_matter_and_scene_pages(self):
        scenes = [
            {"Escena": "1", "Encabezado de escena": "INT. CASA"},
            {"Escena": "2", "Encabezado de escena": "EXT. CALLE"},
        ]
        include = {"cover": True, "summary": True, "catalogs": True, "separators": True}
        pdf = _export_bytes("Hoja de Breakdown", scenes, "PDF", include)
        # Cover + summary + catalog index + six catalog title/data pairs +
        # two scene documents, each preceded by its own editorial separator.
        expected = 2 + 1 + (len(DOCUMENT_DEFINITIONS) - 1) * 2 + 2 + 2
        self.assertEqual(len(PdfReader(BytesIO(pdf)).pages), expected)

    def test_every_scene_including_first_has_editorial_separator(self):
        scenes = [
            {"Escena": "1", "Encabezado de escena": "INT. CASA"},
            {"Escena": "2", "Encabezado de escena": "EXT. CALLE"},
        ]
        pdf = _export_bytes(
            "Hoja de Breakdown", scenes, "PDF",
            {"cover": False, "summary": False, "catalogs": False, "separators": True},
        )
        pages = PdfReader(BytesIO(pdf)).pages
        self.assertEqual(len(pages), 4)
        self.assertIn("ESCENA 1", pages[0].extract_text())
        self.assertIn("ESCENA 2", pages[2].extract_text())

    def test_document_definitions_expose_generators_and_dynamic_options(self):
        self.assertEqual(set(DOCUMENTS_BY_LABEL), {
            "Hoja de Breakdown", "Documento de Cast",
            "Documento de Props y Utilería", "Documento de Vestuario y Maquillaje",
            "Documento de VFX / Efectos Prácticos / Sonido",
            "Documento de Extras, Vehículos y Animales", "Documento de Producción",
        })
        for definition in DOCUMENT_DEFINITIONS:
            self.assertTrue(definition.preview_generator)
            self.assertTrue(definition.pdf_generator)
            self.assertTrue(definition.excel_generator)
            self.assertEqual(definition.header, "cineplan_header")
            self.assertEqual(definition.footer, "cineplan_footer")
            self.assertEqual(definition.editorial_template, "cineplan_editorial")
            self.assertTrue(definition.description)
            self.assertTrue(definition.color.startswith("#"))
            self.assertTrue(callable(definition.scene_color_resolver))

    def test_professional_filenames_follow_document_and_scope(self):
        breakdown = DOCUMENTS_BY_LABEL["Hoja de Breakdown"]
        cast = DOCUMENTS_BY_LABEL["Documento de Cast"]
        self.assertEqual(
            export_filename("Mi Proyecto", breakdown, "Escena actual", [12], "pdf"),
            "Mi Proyecto - Breakdown Escena 12.pdf",
        )
        self.assertEqual(
            export_filename("Mi Proyecto", breakdown, "Escenas seleccionadas", [1, 2, 3], "xlsx"),
            "Mi Proyecto - Breakdown Escenas 1-3.xlsx",
        )
        self.assertEqual(
            export_filename("Mi Proyecto", breakdown, "Proyecto completo", [1, 2], "pdf"),
            "Mi Proyecto - Breakdown Completo.pdf",
        )
        self.assertEqual(
            export_filename("Mi Proyecto", cast, "Proyecto completo", [1], "pdf"),
            "Mi Proyecto - Reparto.pdf",
        )

    def test_production_instruction_uses_saved_and_legacy_keys(self):
        st = __import__("streamlit")
        st.session_state.breakdown_production_notes_data = {
            "1": {"notas_produccion": pd.DataFrame([{
                "Departamento": "Arte", "Instrucción": "Preparar versión hero",
                "Observaciones": "Confirmar continuidad",
            }])},
            "2": {"notas_produccion": pd.DataFrame([{
                "Departamento": "Foto", "Instruccion": "Usar difusión completa",
                "Observaciones": "Prueba previa",
            }])},
        }
        scenes = [{"Escena": "1"}, {"Escena": "2"}]
        document = "Documento de Producción"
        selected = _export_dataframe(document, scenes, {
            "fields": ("Departamento", "Instrucción", "Observaciones"),
        })
        self.assertEqual(list(selected.columns), ["Departamento", "Instrucción", "Observaciones"])
        self.assertEqual(selected["Instrucción"].tolist(), [
            "Preparar versión hero", "Usar difusión completa",
        ])
        without_instruction = _export_dataframe(document, scenes, {
            "fields": ("Departamento", "Observaciones"),
        })
        self.assertNotIn("Instrucción", without_instruction.columns)

    def test_pdf_widths_fill_page_and_prioritize_long_text(self):
        columns = ("ID", "Departamento", "Instrucción", "Estado", "Escenas")
        widths = pdf_column_widths(columns, 760)
        weights = column_weights(columns)
        self.assertAlmostEqual(sum(widths), 760)
        self.assertGreater(widths[2], widths[0])
        self.assertGreater(weights[2], weights[3])

    def test_editorial_grid_normalizes_page_and_table_metrics(self):
        self.assertEqual(GRID.pdf_content_top, THEME.header_height + THEME.margin + 14)
        self.assertEqual(GRID.pdf_content_bottom, THEME.footer_height + THEME.margin)
        self.assertGreater(GRID.table_header_height, 0)
        self.assertEqual(GRID.table_padding_x, 5)

    def test_stripboard_color_is_shared_by_future_document_definitions(self):
        identity = scene_color_identity({"Color stripboard": "Azul"})
        legacy = scene_color_identity({"color_stripboard": "Morado"})
        neutral = scene_color_identity({})
        self.assertEqual(identity.hex, "#b7dcff")
        self.assertEqual(legacy.hex, "#d1c4e9")
        self.assertTrue(identity.assigned)
        self.assertFalse(neutral.assigned)
        definition_identity = DOCUMENTS_BY_LABEL["Hoja de Breakdown"].scene_color(
            {"Color stripboard": "Azul"}
        )
        self.assertEqual(definition_identity, identity)

    def test_preview_and_pdf_breakdown_accents_share_resolved_color(self):
        identity = scene_color_identity({"Color stripboard": "Verde"})
        self.assertIn(identity.hex, html_scene_accent(identity.hex, identity.label))
        commands = pdf_scene_accent_commands(identity.hex)
        self.assertEqual(commands[0][0], "LINEABOVE")
        self.assertEqual(commands[0][3], 5)
        self.assertEqual(commands[0][4].hexval().lower(), "0xc8e6c9")

    def test_scene_separator_receives_assigned_color_and_neutral_fallback(self):
        with patch("modules.breakdown.export_page.editorial_pdf", return_value=b"PDF") as renderer:
            self.assertEqual(_scene_separator_pdf({"Escena": "4", "Color stripboard": "Rosa"}), b"PDF")
            self.assertEqual(renderer.call_args.kwargs["color"], "#f8bbd0")
            self.assertTrue(renderer.call_args.kwargs["full_color_bar"])
        with patch("modules.breakdown.export_page.editorial_pdf", return_value=b"PDF") as renderer:
            _scene_separator_pdf({"Escena": "5"})
            self.assertEqual(renderer.call_args.kwargs["color"], "#d1d5db")

    def test_column_balance_uses_visible_content_without_wasting_width(self):
        columns = ("ID", "Nombre", "Observaciones")
        rows = [{"ID": "1", "Nombre": "Luz", "Observaciones": "Texto editorial largo " * 8}]
        widths = pdf_column_widths(columns, 760, rows)
        self.assertAlmostEqual(sum(widths), 760)
        self.assertGreater(widths[2], widths[1])
        self.assertGreater(widths[1], widths[0])

    def test_shared_preview_and_pdf_components_use_same_visual_language(self):
        header = html_header("NOTAS DE PRODUCCIÓN", (("Proyecto", "Luna"),))
        footer = html_footer("22/07/2026")
        self.assertIn("#09090b", header)
        self.assertIn("CINEPLAN SCHEDULER", header)
        self.assertIn("CinePlan Scheduler", footer)
        frame = pd.DataFrame([{"Departamento": "Arte", "Instrucción": "Preparar hero"}])
        pdf = dataframe_pdf(
            frame, "NOTAS DE PRODUCCIÓN",
            pdf_column_widths(frame.columns, THEME.usable_width), project="Luna", version="3",
        ).getvalue()
        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
        self.assertIn("CINEPLAN SCHEDULER", text)
        self.assertIn("NOTAS DE PRODUCCIÓN", text)
        self.assertIn("Preparar hero", text)

    def test_editorial_template_has_shared_header_footer_and_metadata(self):
        pdf = editorial_pdf(
            "ESCENA 12", "INT. COCINA - NOCHE", "Separador de escena",
            (("Tiempo", "Noche"), ("Octavos", "2/8")), project="Luna", version="3",
        )
        text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
        self.assertIn("CINEPLAN SCHEDULER", text)
        self.assertIn("ESCENA 12", text)
        self.assertIn("Página 1", text)
        self.assertIn("OCTAVOS", text)

    def test_production_excel_wraps_complete_instruction(self):
        st = __import__("streamlit")
        instruction = "Instrucción extensa que debe conservarse completa y ajustarse dentro de la celda."
        st.session_state.breakdown_production_notes_data = {
            number: {"notas_produccion": pd.DataFrame([{
                "Departamento": "Producción", "Instrucción": instruction,
                "Observaciones": "Sin observaciones",
            }])} for number in ("1", "2")
        }
        data = _export_bytes(
            "Documento de Producción", [{"Escena": "1"}, {"Escena": "2"}], "Excel", {},
            {"fields": ("Departamento", "Instrucción", "Observaciones")},
        )
        workbook = load_workbook(BytesIO(data))
        worksheet = workbook["Escena 1"]
        headers = [cell.value for cell in worksheet[1]]
        instruction_column = headers.index("Instrucción") + 1
        self.assertEqual(worksheet.cell(2, instruction_column).value, instruction)
        self.assertTrue(worksheet.cell(2, instruction_column).alignment.wrap_text)
        self.assertGreaterEqual(worksheet.column_dimensions[worksheet.cell(1, instruction_column).column_letter].width, 28)

    def test_zoom_and_fit_update_preview_scale_only(self):
        app = AppTest.from_string("""
from modules.breakdown.export_page import _render_zoom_controls
_render_zoom_controls()
""").run(timeout=20)
        app.button(key="export_zoom_in").click().run(timeout=20)
        self.assertEqual(app.session_state["export_preview_scale"], 1.1)
        app.button(key="export_fit").click().run(timeout=20)
        self.assertEqual(app.session_state["export_preview_scale"], 1.0)


if __name__ == "__main__":
    unittest.main()
